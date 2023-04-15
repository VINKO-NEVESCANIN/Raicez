import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#Poner la ruta del acrivo excel a consultar
archivo_excel = pd.read_excel(r'C:\Users\VINKO\Documents\GitHub\Raicez\RAICEZ\Excel\Temperatura_control.xlsx')

#archivo_excel[['TtarRC_Avg(1)', 'TtarRC_Avg(2)','TtarRC_Avg(3)', 'TtarRC_Avg(4)', 'TtarRC_Avg(5)', 'TtarRC_Avg(6)', 'TtarRC_Avg(7)', 'TtarRC_Avg(8)', 'TtarHC_Avg(1)', 'TtarHC_Avg(2)', 'TtarHC_Avg(3)', 'TtarHC_Avg(4)', 'TtarHC_Avg(5)', 'TtarHC_Avg(6)', 'TtarHC_Avg(7)', 'TtarHC_Avg(8)']]



#Columnas tomas encuenta para los datos a registrar
datos = {'TtarRC_Avg(1)', 'TtarRC_Avg(2)', 'TtarRC_Avg(3)', 'TtarRC_Avg(4)', 'TtarRC_Avg(5)', 'TtarRC_Avg(6)', 'TtarRC_Avg(7)', 'TtarRC_Avg(8)'}

Tabla_pivote = archivo_excel.pivot_table(
    index='Date', columns='TIMESTAMP', values= datos, aggfunc='sum').round(0)

print(Tabla_pivote)

Tabla_pivote.to_excel('Temperaturas_2022.xlsx',startrow=3,sheet_name='Report')

wb = load_workbook('Temperaturas_2022.xlsx')
pestana = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

print(min_col)
print(max_col)
print(min_fila)
print(max_fila)


#Grafico
# barchart = BarChart()

# data = Reference(pestana, min_col=min_col+1, max_col=max_col,min_row=min_fila,max_row=max_fila)
# categorias = Reference(pestana, min_col=min_col+1, max_col=max_col,min_row=min_fila, max_row=max_fila)

# barchart.add_data(data, tittles_from_data=True)
# barchart.set_categories(categorias)

# pestana.add_chart(barchart, 'E72')
# barcahrt.title = 'Ventanas'
# barchart.style = 2 #5

#pestana['B8'] = '=SUM(B6:B7)'
#pestana['B8'].style = 'currency'

#pestana['C6'] = '=SUM(B6:B7)'
#pestana['C7'].style = 'currency'

# abecedario = list(string.ascii_uppercase)
# #print(abecedario[0:max_col])
# abecedario_excel = abecedario[0:max_col]

# for i in abecedario_excel:
#     if i!='A':
#         pestana[f'{i}{max_fila}'] = f'SUM({i}6:{i}7'
#         pestana[f'{i}8'].style = 'Currency'
 
# pestana[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'       
# #pestana['B8'] = '=SUM(B6:B7)'        

wb.save('Temperaturas_2022.xlsx')

#TIMESTAMP