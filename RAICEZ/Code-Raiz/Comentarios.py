import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

def filtrar_datos(archivo, columnas_seleccionadas, rangos_columnas):
    df = pd.read_excel(archivo)
    conditions = []  # Para aplicar formato condicional

    # Aplicar filtro y preparar condiciones para formato
    for columna, (caja_min, caja_max) in zip(columnas_seleccionadas, rangos_columnas):
        rango_min = float(caja_min.get() or 0)
        rango_max = float(caja_max.get() or float('inf'))

        # Convertir la columna a números flotantes y preparar condiciones
        df[columna] = pd.to_numeric(df[columna], errors='coerce')
        condition = (df[columna] < rango_min) | (df[columna] > rango_max)
        conditions.append((columna, condition))

    # Guardar los datos filtrados
    nombre_archivo = os.path.splitext(os.path.basename(archivo))[0]
    archivo_salida = f'{nombre_archivo}_datos_filtrados.xlsx'
    df.to_excel(archivo_salida, index=False)

    # Aplicar formato condicional con openpyxl
    wb = load_workbook(archivo_salida)
    ws = wb.active
    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Rojo

    for col_idx, column in enumerate(df.columns, 1):
        for condition in conditions:
            if column == condition[0]:
                for row_idx in range(2, ws.max_row + 1):  # Desde la segunda fila hasta la última
                    if condition[1][row_idx - 2]:  # condition[1] es la serie de booleanos
                        ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(archivo_salida)
    print(f"Los datos filtrados se han guardado en '{archivo_salida}'")
    # Abrir automáticamente el archivo Excel generado
    os.startfile(archivo_salida)

    # Procesar el DataFrame filtrado...





# Crear la ventana principal
ventana = tk.Tk()
ventana.title('Selección de datos de Excel')

# Etiqueta y campo de entrada para la ruta del archivo
label_ruta_archivo = tk.Label(ventana, text='Archivo de Excel:')
label_ruta_archivo.grid(row=0, column=0)
entry_ruta_archivo = tk.Entry(ventana, width=40)
entry_ruta_archivo.grid(row=0, column=1)
button_cargar_archivo = tk.Button(ventana, text='Cargar', command=cargar_archivo)
button_cargar_archivo.grid(row=0, column=2)

# Etiquetas y campos de entrada para la selección de datos
label_columnas = tk.Label(ventana, text='Columnas seleccionadas (separadas por coma):')
label_columnas.grid(row=1, column=0)
entry_columnas = tk.Entry(ventana, width=40)
entry_columnas.grid(row=1, column=1, columnspan=2)

# Botón para procesar los datos
button_procesar = tk.Button(ventana, text='Procesar', command=procesar_datos)
button_procesar.grid(row=2, column=0, columnspan=3)

# Iniciar el bucle principal de la ventana
ventana.mainloop()
