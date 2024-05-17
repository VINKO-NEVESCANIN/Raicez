import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.simpledialog import askstring
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

epsilon = 1.001  # Valor pequeño para ajustar los límites

def cargar_archivo():
    ruta_archivo = filedialog.askopenfilename(filetypes=[('Archivos Excel', '*.xlsx')])
    if ruta_archivo:
        entry_ruta_archivo.delete(0, tk.END)
        entry_ruta_archivo.insert(tk.END, ruta_archivo)

def procesar_datos():
    archivo = entry_ruta_archivo.get()
    if archivo:
        columnas_seleccionadas = entry_columnas.get().split(',')
        if columnas_seleccionadas:
            if askstring("Input", "¿Desea ingresar los mismos valores para todas las columnas? (s/n)").lower() == 's':
                valor_minimo = float(askstring("Input", "Ingrese el valor mínimo para todas las columnas:"))
                valor_maximo = float(askstring("Input", "Ingrese el valor máximo para todas las columnas:"))
                filtrar_datos(archivo, columnas_seleccionadas, valor_minimo, valor_maximo)
            else:
                filtrar_datos(archivo, columnas_seleccionadas)
        else:
            messagebox.showinfo("Información", "Por favor, introduce las columnas.")
    else:
        messagebox.showinfo("Información", "Por favor, carga un archivo.")

def filtrar_datos(archivo, columnas_seleccionadas, valor_minimo=None, valor_maximo=None):
    df = pd.read_excel(archivo)
    conditions = []

    if valor_minimo is not None and valor_maximo is not None:
        for columna in columnas_seleccionadas:
            df[columna] = pd.to_numeric(df[columna], errors='coerce')
            condition = (df[columna] < valor_minimo) | (df[columna] > valor_maximo + epsilon)
            conditions.append((columna, condition))
            print(f"Límites para la columna {columna}: Mínimo = {valor_minimo}, Máximo = {valor_maximo}")
            print(f"Valores fuera de rango para la columna {columna}:")
            print(df[columna][condition])
    else:
        for columna in columnas_seleccionadas:
            rango_min = float(askstring("Input", f"Introduce el mínimo para {columna}:"))
            rango_max = float(askstring("Input", f"Introduce el máximo para {columna}:"))
            df[columna] = pd.to_numeric(df[columna], errors='coerce')
            condition = (df[columna] < rango_min) | (df[columna] > rango_max + epsilon)
            conditions.append((columna, condition))
            print(f"Límites para la columna {columna}: Mínimo = {rango_min}, Máximo = {rango_max}")
            print(f"Valores fuera de rango para la columna {columna}:")
            print(df[columna][condition])

    nombre_archivo = os.path.splitext(os.path.basename(archivo))[0]
    archivo_salida = f'{nombre_archivo}_datos_filtrados.xlsx'
    df.to_excel(archivo_salida, index=False)

    wb = load_workbook(archivo_salida)
    ws = wb.active
    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for col_idx, column in enumerate(df.columns, 1):
        if column in columnas_seleccionadas:
            idx = columnas_seleccionadas.index(column)
            for row_idx, valor in enumerate(df[column], 2):
                if conditions[idx][1][row_idx - 2]:
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(archivo_salida)
    print(f"Los datos filtrados se han guardado en '{archivo_salida}'")
    os.startfile(archivo_salida)

    # Generar gráfico para las columnas seleccionadas
    generar_grafico(df, columnas_seleccionadas, valor_minimo, valor_maximo)

    print("Columnas disponibles en el DataFrame:", df.columns)

def generar_grafico(df, columnas_seleccionadas, valor_minimo=None, valor_maximo=None):
    fig, ax = plt.subplots(figsize=(10, 6))
    
    porcentajes_error = []
    for columna in columnas_seleccionadas:
        if columna in df.columns:
            if valor_minimo is not None and valor_maximo is not None:
                condition = (df[columna] < valor_minimo) | (df[columna] > valor_maximo)
                porcentaje = df[condition][columna].count() / df[columna].count() * 100
            else:
                porcentaje = 0
            
            if porcentaje == 0:
                print(f"No hay datos fuera de rango para la columna {columna}. No se puede generar la gráfica.")
            
            porcentajes_error.append(porcentaje)
        else:
            print(f"Columna {columna} no encontrada en el DataFrame.")
            porcentajes_error.append(0)
    
    ax.bar(columnas_seleccionadas, porcentajes_error, color='skyblue')
    ax.set_ylabel('Porcentaje de Error')
    ax.set_title('Porcentaje de Error para las Columnas Seleccionadas')
    ax.set_xlabel('Columnas')

    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

ventana = tk.Tk()
ventana.title('Selección de datos de Excel')

label_ruta_archivo = tk.Label(ventana, text='Archivo de Excel:')
label_ruta_archivo.grid(row=0, column=0)
entry_ruta_archivo = tk.Entry(ventana, width=40)
entry_ruta_archivo.grid(row=0, column=1)
button_cargar_archivo = tk.Button(ventana, text='Cargar', command=cargar_archivo)
button_cargar_archivo.grid(row=0, column=2)

label_columnas = tk.Label(ventana, text='Columnas seleccionadas (separadas por coma):')
label_columnas.grid(row=1, column=0)
entry_columnas = tk.Entry(ventana, width=40)
entry_columnas.grid(row=1, column=1, columnspan=2)

button_procesar = tk.Button(ventana, text='Procesar', command=procesar_datos)
button_procesar.grid(row=2, column=0, columnspan=3)

ventana.mainloop()
