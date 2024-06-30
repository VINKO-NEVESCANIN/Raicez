import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import schedule
import time
from threading import Thread
from datetime import datetime

epsilon = 1.001  # Valor pequeño para ajustar los límites

# Variables globales para la automatización
archivo_global = None
columnas_seleccionadas_global = None
valor_minimo_global = None
valor_maximo_global = None
ruta_guardado_global = None

def cargar_archivo():
    ruta_archivo = filedialog.askopenfilename(filetypes=[('Archivos Excel', '*.xlsx')])
    if ruta_archivo:
        entry_ruta_archivo.delete(0, tk.END)
        entry_ruta_archivo.insert(tk.END, ruta_archivo)

def seleccionar_guardado():
    global ruta_guardado_global
    ruta_guardado_global = filedialog.askdirectory()
    if ruta_guardado_global:
        entry_ruta_guardado.delete(0, tk.END)
        entry_ruta_guardado.insert(tk.END, ruta_guardado_global)

def procesar_datos():
    global archivo_global, columnas_seleccionadas_global, valor_minimo_global, valor_maximo_global
    
    archivo = entry_ruta_archivo.get()
    if archivo:
        columnas_seleccionadas = [col.strip() for col in entry_columnas.get().split(',')]
        if columnas_seleccionadas:
            respuesta = messagebox.askyesno("Input", "¿Desea ingresar los mismos valores para todas las columnas?")
            if respuesta:  # Si el usuario selecciona "Sí"
                try:
                    valor_minimo = float(simpledialog.askstring("Input", "Ingrese el valor mínimo para todas las columnas:"))
                    valor_maximo = float(simpledialog.askstring("Input", "Ingrese el valor máximo para todas las columnas:"))
                    archivo_global = archivo
                    columnas_seleccionadas_global = columnas_seleccionadas
                    valor_minimo_global = valor_minimo
                    valor_maximo_global = valor_maximo
                    seleccionar_guardado()
                    filtrar_datos(archivo, columnas_seleccionadas, valor_minimo, valor_maximo)
                except (TypeError, ValueError):
                    messagebox.showerror("Error", "Valor mínimo o máximo no válido. Operación cancelada.")
            else:  # Si el usuario selecciona "No"
                try:
                    archivo_global = archivo
                    columnas_seleccionadas_global = columnas_seleccionadas
                    seleccionar_guardado()
                    filtrar_datos(archivo, columnas_seleccionadas)
                except (TypeError, ValueError):
                    messagebox.showerror("Error", "Valores no válidos. Operación cancelada.")
        else:
            messagebox.showinfo("Información", "Por favor, introduce las columnas.")
    else:
        messagebox.showinfo("Información", "Por favor, carga un archivo.")

def filtrar_datos(archivo, columnas_seleccionadas, valor_minimo=None, valor_maximo=None):
    df = pd.read_excel(archivo)
    conditions = []
    columnas_no_encontradas = []
    datos_fuera_de_rango = False

    if valor_minimo is not None and valor_maximo is not None:
        for columna in columnas_seleccionadas:
            if columna in df.columns:
                df[columna] = pd.to_numeric(df[columna], errors='coerce')
                condition = (df[columna] < valor_minimo) | (df[columna] > valor_maximo + epsilon)
                conditions.append((columna, condition))
                if df[columna][condition].count() > 0:
                    datos_fuera_de_rango = True
                print(f"Límites para la columna {columna}: Mínimo = {valor_minimo}, Máximo = {valor_maximo}")
                print(f"Valores fuera de rango para la columna {columna}:")
                print(df[columna][condition])
            else:
                columnas_no_encontradas.append(columna)
    else:
        for columna in columnas_seleccionadas:
            if columna in df.columns:
                try:
                    rango_min = float(simpledialog.askstring("Input", f"Introduce el mínimo para {columna}:"))
                    rango_max = float(simpledialog.askstring("Input", f"Introduce el máximo para {columna}:"))
                    df[columna] = pd.to_numeric(df[columna], errors='coerce')
                    condition = (df[columna] < rango_min) | (df[columna] > rango_max + epsilon)
                    conditions.append((columna, condition))
                    if df[columna][condition].count() > 0:
                        datos_fuera_de_rango = True
                    print(f"Límites para la columna {columna}: Mínimo = {rango_min}, Máximo = {rango_max}")
                    print(f"Valores fuera de rango para la columna {columna}:")
                    print(df[columna][condition])
                except (TypeError, ValueError):
                    messagebox.showerror("Error", f"Valores no válidos para la columna {columna}. Operación cancelada.")
                    return
            else:
                columnas_no_encontradas.append(columna)

    if columnas_no_encontradas:
        messagebox.showinfo("Información", f"Columnas no encontradas en el archivo: {', '.join(columnas_no_encontradas)}")

    if not datos_fuera_de_rango:
        messagebox.showinfo("Información", "No hay datos fuera de rango para las columnas seleccionadas.")
        return

    nombre_archivo = os.path.splitext(os.path.basename(archivo))[0]
    archivo_salida = os.path.join(ruta_guardado_global, f'{nombre_archivo}_datos_filtrados.xlsx')
    df.to_excel(archivo_salida, index=False)

    wb = load_workbook(archivo_salida)
    ws = wb.active
    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for col_idx, column in enumerate(df.columns, 1):
        if column in columnas_seleccionadas:
            idx = columnas_seleccionadas.index(column)
            for row_idx, valor in enumerate(df[column], 2):
                if conditions[idx][1][row_idx - 2]:
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(archivo_salida)
    print(f"Los datos filtrados se han guardado en '{archivo_salida}'")
    os.startfile(archivo_salida)

    # Generar gráfico para las columnas seleccionadas
    generar_grafico(df, columnas_seleccionadas, conditions)

    print("Columnas disponibles en el DataFrame:", df.columns)

def generar_grafico(df, columnas_seleccionadas, conditions):
    fig, ax = plt.subplots(figsize=(10, 6))
    
    porcentajes_error = []
    for idx, columna in enumerate(columnas_seleccionadas):
        if columna in df.columns:
            condition = conditions[idx][1]
            porcentaje = df[condition].shape[0] / df[columna].shape[0] * 100
            porcentajes_error.append(porcentaje)
        else:
            print(f"Columna {columna} no encontrada en el DataFrame.")
            porcentajes_error.append(0)
    
    ax.bar(columnas_seleccionadas, porcentajes_error, color='skyblue')
    ax.set_ylabel('Porcentaje de Error')
    ax.set_title('Porcentaje de Error para las Columnas Seleccionadas')
    ax.set_xlabel('Columnas')
    ax.set_yticks(range(0, 101, 10))  # Ajustar el eje y para mostrar más valores de porcentaje

    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

def tarea_automatica():
    if archivo_global and columnas_seleccionadas_global:
        try:
            if valor_minimo_global is not None and valor_maximo_global is not None:
                filtrar_datos(archivo_global, columnas_seleccionadas_global, valor_minimo_global, valor_maximo_global)
            else:
                filtrar_datos(archivo_global, columnas_seleccionadas_global)
        except Exception as e:
            print(f"Error en la tarea automática: {e}")

def configurar_automatizacion():
    # Crear una nueva ventana para la configuración de la automatización
    ventana_autom = tk.Toplevel()
    ventana_autom.title('Configurar Automatización')

    dias_vars = {}
    dias = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']
    
    for i, dia in enumerate(dias):
        var = tk.BooleanVar()
        chk = tk.Checkbutton(ventana_autom, text=dia.capitalize(), variable=var)
        chk.grid(row=i, column=0, sticky='w')
        dias_vars[dia] = var
    
    label_hora = tk.Label(ventana_autom, text='Hora (HH:MM, 24h):')
    label_hora.grid(row=7, column=0)
    entry_hora = tk.Entry(ventana_autom)
    entry_hora.grid(row=7, column=1)
    
    def guardar_configuracion():
        dias_seleccionados = [dia for dia, var in dias_vars.items() if var.get()]
        hora = entry_hora.get()
        
        try:
            datetime.strptime(hora, '%H:%M')  # Validar el formato de la hora
            if dias_seleccionados:
                configurar_schedule(dias_seleccionados, hora)
                ventana_autom.destroy()
            else:
                messagebox.showerror("Error", "Seleccione al menos un día para la automatización.")
        except ValueError:
            messagebox.showerror("Error", "Hora no válida. Use el formato HH:MM.")
    
    button_guardar = tk.Button(ventana_autom, text='Guardar', command=guardar_configuracion)
    button_guardar.grid(row=8, column=0, columnspan=2)
    
def configurar_schedule(dias, hora):
    if ruta_guardado_global:
        try:
            def tarea_automatica():
                if archivo_global and columnas_seleccionadas_global:
                    try:
                        if valor_minimo_global is not None and valor_maximo_global is not None:
                            filtrar_datos(archivo_global, columnas_seleccionadas_global, valor_minimo_global, valor_maximo_global)
                        else:
                            filtrar_datos(archivo_global, columnas_seleccionadas_global)
                    except Exception as e:
                        print(f"Error en la tarea automática: {e}")

            for dia in dias:
                dia_schedule = getattr(schedule.every(), dia)
                dia_schedule.at(hora).do(tarea_automatica)
            
            messagebox.showinfo("Automatización", "La automatización ha sido configurada.")
            
            # Iniciar el hilo para ejecutar `schedule`
            def run_scheduler():
                while True:
                    schedule.run_pending()
                    time.sleep(1)
            
            scheduler_thread = Thread(target=run_scheduler, daemon=True)
            scheduler_thread.start()
        except Exception as e:
            messagebox.showerror("Error", f"Error al iniciar la automatización: {e}")
    else:
        messagebox.showerror("Error", "Seleccione una carpeta para guardar los archivos antes de iniciar la automatización.")

ventana = tk.Tk()
ventana.title('Selección de datos de Excel')

label_ruta_archivo = tk.Label(ventana, text='Archivo de Excel:')
label_ruta_archivo.grid(row=0, column=0)
entry_ruta_archivo = tk.Entry(ventana, width=40)
entry_ruta_archivo.grid(row=0, column=1)
button_cargar_archivo = tk.Button(ventana, text='Cargar', command=cargar_archivo)
button_cargar_archivo.grid(row=0, column=2)

label_columnas = tk.Label(ventana, text='Columnas seleccionadas (separadas por coma):')
label_columnas.grid(row=1, column=0)
entry_columnas = tk.Entry(ventana, width=40)
entry_columnas.grid(row=1, column=1, columnspan=2)

label_ruta_guardado = tk.Label(ventana, text='Carpeta para guardar:')
label_ruta_guardado.grid(row=2, column=0)
entry_ruta_guardado = tk.Entry(ventana, width=40)
entry_ruta_guardado.grid(row=2, column=1)
button_seleccionar_guardado = tk.Button(ventana, text='Seleccionar', command=seleccionar_guardado)
button_seleccionar_guardado.grid(row=2, column=2)

button_procesar = tk.Button(ventana, text='Procesar', command=procesar_datos)
button_procesar.grid(row=3, column=0, columnspan=3)

button_automatizar = tk.Button(ventana, text='Configurar Automatización', command=configurar_automatizacion)
button_automatizar.grid(row=4, column=0, columnspan=3)

ventana.mainloop()
